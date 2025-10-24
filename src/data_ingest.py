from __future__ import annotations
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Tuple, Iterable
import openpyxl
import re
from datetime import datetime, date, time, timedelta

SPANISH_MONTHS = {
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}


def parse_spanish_number(text: str) -> float:
    t = (text or '').strip()
    t = t.replace('.', '')
    t = t.replace(',', '.')
    return float(t)


def parse_spanish_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    if not isinstance(value, str):
        return None
    s = value.strip()
    s = s.replace('CET', '').replace('CEST', '').strip()
    # Examples: '22 ene 2024 00:00:00'
    m = re.match(r"^(\d{1,2})\s+(\w{3})\s+(\d{4})\s+(\d{2}:\d{2}:\d{2})$", s, re.IGNORECASE)
    if m:
        day = int(m.group(1))
        mon_txt = m.group(2).lower()
        year = int(m.group(3))
        hhmmss = m.group(4)
        month = SPANISH_MONTHS.get(mon_txt)
        if not month:
            return None
        hh, mm, ss = map(int, hhmmss.split(':'))
        try:
            return datetime(year, month, day, hh, mm, ss)
        except ValueError:
            return None
    # Fallback: try ISO or common formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


@dataclass
class HDDRecord:
    date: datetime
    hdd: float


def load_hdd_series(hdd_dir: Path) -> List[HDDRecord]:
    records: List[HDDRecord] = []
    for path in sorted(hdd_dir.glob('*.xlsx')):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if row is None or all(v is None for v in row):
                    continue
                date_val, hdd_val = row[0], row[1]
                dt = parse_spanish_datetime(date_val)
                if dt is None:
                    continue
                try:
                    val = float(hdd_val)
                except Exception:
                    continue
                records.append(HDDRecord(date=dt, hdd=val))
        finally:
            wb.close()
    return records


def load_energy_series(prod_dir: Path) -> Dict[str, List[Tuple[datetime, float]]]:
    series: Dict[str, List[Tuple[datetime, float]]] = {}
    # Filenames like m218807.xlsx or m218820-1.xlsx
    for path in sorted(prod_dir.glob('*.xlsx')):
        m = re.search(r'm(\d+)', path.stem, flags=re.IGNORECASE)
        if not m:
            continue
        inst_code = f"M{m.group(1)}"
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if row is None or all(v is None for v in row):
                    continue
                ts_val, energy_val = row[0], row[1]
                dt = parse_spanish_datetime(ts_val)
                if dt is None:
                    continue
                if isinstance(energy_val, str):
                    try:
                        val = parse_spanish_number(energy_val)
                    except Exception:
                        continue
                else:
                    try:
                        val = float(energy_val)
                    except Exception:
                        continue
                series.setdefault(inst_code, []).append((dt, val))
        finally:
            wb.close()
    for k in series:
        series[k].sort(key=lambda x: x[0])
    return series


def load_biomass_deliveries(path: Path) -> List[Tuple[str, datetime, float]]:
    deliveries: List[Tuple[str, datetime, float]] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb['Descargas'] if 'Descargas' in wb.sheetnames else wb.worksheets[0]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row is None or all(v is None for v in row):
                continue
            inst, fecha, cantidad = row[0], row[1], row[2]
            if not isinstance(inst, str):
                continue
            dt = parse_spanish_datetime(fecha)
            if dt is None:
                continue
            if isinstance(cantidad, str):
                try:
                    qty = parse_spanish_number(cantidad)
                except Exception:
                    continue
            else:
                try:
                    qty = float(cantidad)
                except Exception:
                    continue
            deliveries.append((inst.strip(), dt, qty))
    finally:
        wb.close()
    return deliveries


def to_daily_increments(series: List[Tuple[datetime, float]]) -> List[Tuple[datetime, float]]:
    if not series:
        return []
    out: List[Tuple[datetime, float]] = []
    prev_day = series[0][0].date()
    day_first_val = series[0][1]
    day_last_val = series[0][1]
    for ts, val in series[1:]:
        d = ts.date()
        if d == prev_day:
            day_last_val = val
            continue
        inc = max(0.0, day_last_val - day_first_val)
        out.append((datetime.combine(prev_day, time()), inc))
        prev_day = d
        day_first_val = val
        day_last_val = val
    # last day increment
    inc = max(0.0, day_last_val - day_first_val)
    out.append((datetime.combine(prev_day, time()), inc))
    return out


def group_by_week(dateseries: Iterable[Tuple[datetime, float]]) -> Dict[date, float]:
    weekly: Dict[date, float] = {}
    for dt, val in dateseries:
        week_start = (dt.date() - timedelta(days=dt.weekday()))
        weekly[week_start] = weekly.get(week_start, 0.0) + float(val)
    return weekly


def series_to_map(dateseries: Iterable[Tuple[datetime, float]]) -> Dict[date, float]:
    m: Dict[date, float] = {}
    for dt, val in dateseries:
        m[dt.date()] = float(val)
    return m

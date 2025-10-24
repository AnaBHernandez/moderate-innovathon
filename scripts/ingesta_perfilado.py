from __future__ import annotations
from pathlib import Path
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple
import openpyxl
import numpy as np
from datetime import datetime

BASE = Path('/workspace/datos_originales')
OUTPUT = Path('/workspace/Entregables')
OUTPUT.mkdir(parents=True, exist_ok=True)

@dataclass
class HDDRecord:
    date: datetime
    hdd: float


def load_hdd_series(hdd_dir: Path) -> List[HDDRecord]:
    records: List[HDDRecord] = []
    for path in sorted(hdd_dir.glob('*.xlsx')):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue  # header
                if row is None or all(v is None for v in row):
                    continue
                date_val, hdd_val = row[0], row[1]
                if isinstance(date_val, datetime) and isinstance(hdd_val, (int, float)):
                    records.append(HDDRecord(date=date_val, hdd=float(hdd_val)))
        finally:
            wb.close()
    return records


def parse_spanish_number(text: str) -> float:
    # values like '153,5' -> 153.5
    t = (text or '').strip()
    t = t.replace('.', '')  # thousands sep
    t = t.replace(',', '.')
    return float(t)


def load_energy_series(prod_dir: Path) -> Dict[str, List[Tuple[datetime, float]]]:
    # Maps installation code to cumulative energy series
    series: Dict[str, List[Tuple[datetime, float]]] = {}
    pattern = re.compile(r'm(\d+)(?:-(\d+))?')
    for path in sorted(prod_dir.glob('*.xlsx')):
        m = pattern.search(path.stem)
        if not m:
            continue
        inst_code = f"M{m.group(1)}"
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue  # header
                if row is None or all(v is None for v in row):
                    continue
                ts_text, energy_text = row[0], row[1]
                if not isinstance(ts_text, str):
                    # allow excel datetime? try fallback
                    if isinstance(ts_text, datetime):
                        ts = ts_text
                    else:
                        continue
                else:
                    # e.g., '22 ene 2024 00:00:00 CET'
                    ts_text = ts_text.replace('CET', '').strip()
                    try:
                        ts = datetime.strptime(ts_text, '%d %b %Y %H:%M:%S')
                    except ValueError:
                        try:
                            ts = datetime.strptime(ts_text, '%d %b %Y %H:%M:%S %Z')
                        except Exception:
                            continue
                if isinstance(energy_text, str):
                    try:
                        val = parse_spanish_number(energy_text)
                    except Exception:
                        continue
                elif isinstance(energy_text, (int, float)):
                    val = float(energy_text)
                else:
                    continue
                series.setdefault(inst_code, []).append((ts, val))
        finally:
            wb.close()
    # sort by datetime
    for k in series:
        series[k].sort(key=lambda x: x[0])
    return series


def load_biomass_deliveries(path: Path) -> List[Tuple[str, datetime, float]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    deliveries: List[Tuple[str, datetime, float]] = []
    try:
        ws = wb['Descargas'] if 'Descargas' in wb.sheetnames else wb.worksheets[0]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row is None or all(v is None for v in row):
                continue
            inst, fecha, cantidad = row[0], row[1], row[2]
            if not isinstance(inst, str) or not isinstance(fecha, datetime):
                continue
            if isinstance(cantidad, str):
                try:
                    qty = parse_spanish_number(cantidad)
                except Exception:
                    continue
            elif isinstance(cantidad, (int, float)):
                qty = float(cantidad)
            else:
                continue
            deliveries.append((inst.strip(), fecha, qty))
    finally:
        wb.close()
    return deliveries


def daily_energy_from_cumulative(series: List[Tuple[datetime, float]]) -> List[Tuple[datetime, float]]:
    # Convert cumulative MW-hr readings to daily increments
    if not series:
        return []
    daily: List[Tuple[datetime, float]] = []
    prev_val = None
    prev_day = None
    for ts, val in series:
        day = ts.date()
        if prev_day is None:
            prev_day = day
            prev_val = val
            continue
        if day != prev_day:
            inc = max(0.0, val - (prev_val if prev_val is not None else val))
            daily.append((datetime.combine(prev_day, datetime.min.time()), inc))
            prev_day = day
            prev_val = val
        else:
            # same day, keep last value
            prev_val = val
    return daily


def main():
    hdd = load_hdd_series(BASE / 'hdd-anual')
    print(f"HDD registros: {len(hdd)} | rango: {min(r.date for r in hdd).date()} .. {max(r.date for r in hdd).date()}")

    energy_series = load_energy_series(BASE / 'produccion-energetica')
    print(f"Instalaciones con energía: {sorted(energy_series.keys())}")
    # producir energía diaria por instalación
    energy_daily: Dict[str, List[Tuple[datetime, float]]] = {}
    for inst, series in energy_series.items():
        energy_daily[inst] = daily_energy_from_cumulative(series)
        print(f"{inst}: {len(energy_daily[inst])} días de energía")

    deliveries = load_biomass_deliveries(BASE / 'consumo-biomasa.xlsx')
    print(f"Descargas de biomasa: {len(deliveries)} registros")
    # Mostrar ejemplo por instalación
    sample = {}
    for inst, fecha, qty in deliveries[:10]:
        sample.setdefault(inst, 0)
        sample[inst] += 1
    print({k: v for k, v in list(sample.items())[:5]})


if __name__ == '__main__':
    main()
